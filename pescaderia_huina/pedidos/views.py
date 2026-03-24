import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.forms import inlineformset_factory
from django.db.models import Sum, Q, F
from django.template.loader import get_template
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from xhtml2pdf import pisa
from decimal import Decimal

from .models import Pedido, DetallePedido
from .forms import PedidoForm, DetallePedidoForm
from ventas.models import Venta
from productos.models import Producto
from proveedores.models import Proveedor

# Fábrica de formularios. Permite crear múltiples Detalles vinculados a 1 Pedido
DetallePedidoFormSet = inlineformset_factory(
    Pedido, 
    DetallePedido, 
    form=DetallePedidoForm, 
    extra=1, 
    can_delete=True
)

# ==========================================
# 1. LISTA DE PEDIDOS
# ==========================================
@login_required
def lista_pedidos(request):
    pedidos = Pedido.objects.annotate(
        cantidad_total_calculada=Sum('detalles__cantidad') 
    ).order_by('-fecha')

    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if query:
        pedidos = pedidos.filter(
            Q(id__icontains=query) | 
            Q(proveedor__nombre_contacto__icontains=query) | 
            Q(proveedor__nit__icontains=query) |
            Q(detalles__producto__nombre__icontains=query)
        ).distinct()

    if fecha_inicio:
        pedidos = pedidos.filter(fecha__date__gte=fecha_inicio) 
    if fecha_fin:
        pedidos = pedidos.filter(fecha__date__lte=fecha_fin)     
        
    # --- EXPORTAR REPORTE A PDF ---
    if request.GET.get('export') == 'pdf':
        if fecha_inicio and fecha_fin: texto_rango = f"Del {fecha_inicio} al {fecha_fin}"
        elif fecha_inicio: texto_rango = f"Desde el {fecha_inicio}"
        elif fecha_fin: texto_rango = f"Hasta el {fecha_fin}"
        else:
            meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            texto_rango = f"{meses[timezone.now().month]} {timezone.now().year}"
        
        context = {
            'pedidos': pedidos,
            'fecha_generacion': timezone.now(),
            'mes_reporte': texto_rango,                   
            'total_pedidos': pedidos.count(),               
            'ordenes_confirmadas': pedidos.filter(estado__in=['REC', 'recibido']).count(),   
            'ordenes_pendientes': pedidos.filter(estado__in=['PEN', 'pendiente']).count(),     
            'ordenes_canceladas': pedidos.filter(estado__in=['CAN', 'cancelado']).count(),     
            'suma_total': pedidos.aggregate(total=Sum('valor_total'))['total'] or 0                      
        }
        
        template = get_template('reporte_mensual_pdf.html') 
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Pedidos_Pescaderia.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err: return HttpResponse('Error al generar PDF', status=500)
        return response

    # --- EXPORTAR REPORTE A EXCEL (DISEÑO PROFESIONAL) ---
    elif request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Pedidos"

        # Paleta de colores compartida
        azul_oscuro = "0A3D62"
        verde = "198754"
        gris_claro = "F2F2F2"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color=azul_oscuro, end_color=azul_oscuro, fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

        # Encabezado Principal
        ws.merge_cells('A1:G1')
        ws['A1'] = 'REPORTE DE ÓRDENES DE PEDIDO — PESCADERÍA HUINA'
        ws['A1'].font = Font(bold=True, color="FFFFFF", size=13)
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 28

        # Estadísticas (Sin decimales)
        inversion_total = pedidos.aggregate(t=Sum('valor_total'))['t'] or Decimal('0')
        ws.merge_cells('A2:C2')
        ws['A2'] = f'Total Pedidos: {pedidos.count()}'
        ws.merge_cells('D2:G2')
        ws['D2'] = f'Inversión Total: ${float(inversion_total):,.0f}'
        for cell in [ws['A2'], ws['D2']]:
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            cell.alignment = center_align
        ws.row_dimensions[2].height = 20

        # Cabecera de Tabla
        headers = ['# ID', 'Fecha', 'Proveedor', 'NIT', 'Uds.', 'Estado', 'Valor Total']
        col_widths = [10, 15, 30, 18, 10, 15, 18]
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        ws.row_dimensions[3].height = 18

        # Datos
        for row_num, p in enumerate(pedidos, 4):
            estado_texto = "Recibido" if p.estado in ['REC', 'recibido'] else "Pendiente" if p.estado in ['PEN', 'pendiente'] else "Cancelado"
            row_data = [
                f"#{p.id}", p.fecha.strftime("%d/%m/%Y") if p.fecha else "N/A", p.proveedor.nombre_contacto if p.proveedor else "N/A",
                p.proveedor.nit if p.proveedor else "N/A", p.cantidad_total_calculada or 0, estado_texto.upper(), float(p.valor_total or 0)
            ]

            bg = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid") if estado_texto == "Cancelado" else (PatternFill(start_color=gris_claro, end_color=gris_claro, fill_type="solid") if row_num % 2 == 0 else None)

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                if bg: cell.fill = bg
                if col in [1, 2, 4, 5, 6]: cell.alignment = center_align
                if col == 7:
                    cell.alignment = right_align
                    cell.number_format = '$#,##0' # Formato modificado para no tener decimales
                if col == 6:
                    if estado_texto == "Cancelado": cell.font = Font(bold=True, color="DC2626")
                    elif estado_texto == "Recibido": cell.font = Font(bold=True, color=verde)

        ws.freeze_panes = 'A4'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Pedidos_Pescaderia.xlsx"'
        wb.save(response)
        return response

    return render(request, 'lista_pedidos.html', {'pedidos': pedidos})


# ==========================================
# 2. CREAR, EDITAR, ELIMINAR PEDIDOS
# ==========================================
@login_required
def crear_pedido(request):
    if request.method == 'POST':
        form_pedido = PedidoForm(request.POST)
        formset = DetallePedidoFormSet(request.POST) 
        
        if form_pedido.is_valid() and formset.is_valid():
            try:
                nuevo_pedido = form_pedido.save(commit=False)
                nuevo_pedido.valor_total = 0 
                nuevo_pedido.save() 
                
                total_general = 0
                detalles = formset.save(commit=False)
                for detalle in detalles:
                    detalle.pedido = nuevo_pedido 
                    factor = 0.5 if detalle.presentacion in ['libras', 'Lib', 'libra'] else 1
                    subtotal = float(detalle.cantidad or 0) * float(detalle.precio_unitario or 0) * factor
                    total_general += subtotal
                    detalle.save()
                
                nuevo_pedido.valor_total = total_general
                nuevo_pedido.save()
                
                messages.success(request, f"Pedido #{nuevo_pedido.id} registrado con éxito.")
                return redirect('pedidos:lista_pedidos')
            except Exception as e:
                messages.error(request, f"Ocurrió un error: {e}")
        else:
            messages.error(request, "Por favor corrige los errores del formulario.")
    else:
        form_pedido = PedidoForm()
        formset = DetallePedidoFormSet()

    return render(request, 'crear_pedido.html', {'form_pedido': form_pedido, 'formset': formset})

@login_required
def editar_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    
    # BLOQUEO DE ESTADO ELIMINADO PARA PERMITIR EDICIÓN LIBRE EN CUALQUIER ESTADO

    if request.method == 'POST':
        form_pedido = PedidoForm(request.POST, instance=pedido)
        formset = DetallePedidoFormSet(request.POST, instance=pedido)

        if form_pedido.is_valid() and formset.is_valid():
            pedido_obj = form_pedido.save(commit=False)
            formset.save()
            
            total_recalculado = 0
            for d in pedido_obj.detalles.all():
                factor = 0.5 if d.presentacion in ['libras', 'Lib', 'libra'] else 1
                total_recalculado += float(d.cantidad or 0) * float(d.precio_unitario or 0) * factor
                
            pedido_obj.valor_total = total_recalculado
            pedido_obj.save()
            
            messages.info(request, f"Pedido #{pedido.id} actualizado correctamente.")
            return redirect('pedidos:lista_pedidos')
    else:
        form_pedido = PedidoForm(instance=pedido)
        formset = DetallePedidoFormSet(instance=pedido)

    return render(request, 'editar_pedido.html', {'form_pedido': form_pedido, 'formset': formset, 'pedido': pedido})

@login_required
def eliminar_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    if request.method == 'POST':
        pedido.delete()
        messages.warning(request, f"El pedido #{id} ha sido eliminado.")
        return redirect('pedidos:lista_pedidos')
    return render(request, 'eliminar_pedido.html', {'pedido': pedido})


# ==========================================
# 3. DETALLE Y UTILIDADES AJAX
# ==========================================
@login_required
def detalle_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    if request.GET.get('export') == 'pdf':
        detalles_con_subtotal = []
        for d in pedido.detalles.all():
            cant = float(d.cantidad or 0)
            prec = float(d.precio_unitario or 0)
            factor = 0.5 if d.presentacion in ['libras', 'Lib', 'libra'] else 1
            d.subtotal_calculado = cant * prec * factor
            detalles_con_subtotal.append(d)
            
        template = get_template('pdf_pedido.html')
        html = template.render({'pedido': pedido, 'detalles': detalles_con_subtotal})
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Pedido_{pedido.id}.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err: return HttpResponse('Error al generar PDF', status=500)
        return response

    return render(request, 'detalle_pedido.html', {'pedido': pedido})

@login_required
def cargar_productos_proveedor(request):
    proveedor_id = request.GET.get('proveedor_id')
    if proveedor_id:
        productos = Producto.objects.filter(proveedor_id=proveedor_id).order_by('nombre')
        data = list(productos.values('id', 'nombre'))
    else:
        data = []
    return JsonResponse(data, safe=False)

@login_required
@require_POST
def cambiar_estado_pedido(request, pedido_id):
    try:
        data = json.loads(request.body)
        nuevo_estado = data.get('estado', '').upper()
        pedido = get_object_or_404(Pedido, id=pedido_id)
        
        # Mapear estados largos a códigos cortos
        mapeo_estados = {
            'RECIBIDO': 'REC', 'REC': 'REC',
            'CANCELADO': 'CAN', 'CAN': 'CAN',
            'PENDIENTE': 'PEN', 'PEN': 'PEN',
        }
        
        estado_normalizado = mapeo_estados.get(nuevo_estado)
        
        if estado_normalizado:
            pedido.estado = estado_normalizado
            pedido.save()
            return JsonResponse({'success': True, 'estado': estado_normalizado})
        else:
            return JsonResponse({'success': False, 'error': 'Estado no válido'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ==========================================
# 4. INVENTARIO DE PEDIDOS
# ==========================================
@login_required
def inventario_pedidos(request):
    productos = Producto.objects.filter(estado=True)
    
    def calcular_stock_y_pedidos(queryset):
        for p in queryset:
            # Usar la propiedad stock del modelo que calcula: 
            # stock = recibido (estado REC) - vendido (estado COMPLETADA o PENDIENTE)
            p.stock_disponible = p.stock
        return queryset
    
    pescados = calcular_stock_y_pedidos(productos.filter(tipo_producto='PE'))
    mariscos = calcular_stock_y_pedidos(productos.filter(tipo_producto='MA'))
    pollos = calcular_stock_y_pedidos(productos.filter(tipo_producto='PO'))
    
    # --- EXPORTAR INVENTARIO PDF ---
    if request.GET.get('export') == 'pdf':
        template = get_template('inventario_pdf.html')
        context = {
            'pescados': pescados, 'mariscos': mariscos, 'pollos': pollos, 'fecha_generacion': timezone.now(),
        }
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Inventario_Pescaderia.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err: return HttpResponse('Hubo un error al generar el PDF', status=500)
        return response
    
    # --- EXPORTAR INVENTARIO EXCEL (DISEÑO UNIFICADO) ---
    elif request.GET.get('export') == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Inventario_Pescaderia.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Colores y estilos del Excel de ventas
        azul_oscuro = "0A3D62"
        verde = "198754"
        gris_claro = "F2F2F2"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color=azul_oscuro, end_color=azul_oscuro, fill_type="solid")
        category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        category_font = Font(bold=True, color="0F3976", size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        # Título
        ws.merge_cells('A1:D1')
        ws['A1'].value = 'INVENTARIO DE PRODUCTOS — PESCADERÍA HUINA'
        ws['A1'].font = Font(bold=True, color="FFFFFF", size=13)
        ws['A1'].fill = PatternFill(start_color=azul_oscuro, end_color=azul_oscuro, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Información de generación
        ws.merge_cells('A2:D2')
        ws['A2'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        ws['A2'].alignment = center_align
        ws.row_dimensions[2].height = 18

        # Encabezados de tabla
        headers = ['Nombre Producto', 'Proveedor', 'Stock Disponible', 'Presentación']
        col_widths = [35, 25, 18, 20]
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        ws.row_dimensions[3].height = 18

        # Función para agregar sección
        def agregar_categoria(titulo, productos_list, start_row):
            # Fila de categoría
            ws.merge_cells(f'A{start_row}:D{start_row}')
            cat_cell = ws[f'A{start_row}']
            cat_cell.value = titulo
            cat_cell.font = category_font
            cat_cell.fill = category_fill
            cat_cell.alignment = center_align
            cat_cell.border = thin_border
            ws.row_dimensions[start_row].height = 16
            
            row_num = start_row + 1
            
            # Datos
            for p in productos_list:
                ws[f'A{row_num}'] = p.nombre
                ws[f'B{row_num}'] = p.proveedor.nombre_contacto if p.proveedor else "N/A"
                ws[f'C{row_num}'] = p.stock_disponible
                ws[f'D{row_num}'] = p.get_tipo_presentacion_display()
                
                # Determinación de color para stock
                if p.stock_disponible <= 0:
                    stock_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    stock_font = Font(bold=True, color="9C0006")
                elif p.stock_disponible <= 10:
                    stock_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    stock_font = Font(bold=True, color="9C6500")
                else:
                    stock_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    stock_font = Font(bold=True, color="006100")
                
                for col in range(1, 5):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = thin_border
                    if col == 3:  # Stock
                        cell.fill = stock_fill
                        cell.font = stock_font
                        cell.alignment = center_align
                    elif col == 1:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                    
                    # Alternancia de color de fondo
                    if row_num % 2 == 0 and col != 3:
                        cell.fill = PatternFill(start_color=gris_claro, end_color=gris_claro, fill_type="solid")
                
                ws.row_dimensions[row_num].height = 15
                row_num += 1
            
            return row_num

        # Agregar todas las categorías
        fila_actual = 4
        fila_actual = agregar_categoria("🐟 PESCADOS", pescados, fila_actual)
        fila_actual += 1
        fila_actual = agregar_categoria("🦐 MARISCOS", mariscos, fila_actual)
        fila_actual += 1
        fila_actual = agregar_categoria("🐔 POLLOS", pollos, fila_actual)
        
        ws.freeze_panes = 'A4'
        wb.save(response)
        return response
    
    context = {'pescados': pescados, 'mariscos': mariscos, 'pollos': pollos}
    return render(request, 'inventario_pedidos.html', context)