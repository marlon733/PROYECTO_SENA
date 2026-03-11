import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils import timezone
from xhtml2pdf import pisa
from .models import Proveedor
from .forms import ProveedorForm

# 1. LISTA DE PROVEEDORES
@login_required
def lista_proveedores(request):
    # Filtramos solo los activos
    proveedores = Proveedor.objects.filter(estado=True).order_by('nombre_contacto')

    # Lógica de Búsqueda
    query = request.GET.get('q')
    if query:
        proveedores = proveedores.filter(
            Q(nombre_contacto__icontains=query) | 
            Q(nit__icontains=query)
        )

    # CORRECCIÓN AQUÍ: Quitamos 'proveedores/' de la ruta
    return render(request, 'lista_proveedores.html', {'proveedores': proveedores})

# 2. AGREGAR PROVEEDOR
@login_required
def agregar_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('proveedores:lista_proveedores')
    else:
        form = ProveedorForm()
    # CORRECCIÓN AQUÍ
    return render(request, 'agregar_proveedores.html', {'form': form})

# 3. DETALLE DEL PROVEEDOR
@login_required
def detalle_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    # CORRECCIÓN AQUÍ
    return render(request, 'detalle_proveedores.html', {'proveedor': proveedor})

# 4. MODIFICAR PROVEEDOR
@login_required
def modificar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('proveedores:lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    
    # CORRECCIÓN AQUÍ
    return render(request, 'modificar_proveedores.html', {'form': form, 'proveedor': proveedor})

# 6. EXPORTAR PROVEEDORES A PDF
@login_required
def export_proveedores_pdf(request):
    proveedores = Proveedor.objects.all().order_by('nombre_contacto')

    total_proveedores = proveedores.count()
    total_activos    = proveedores.filter(estado=True).count()
    total_inactivos  = proveedores.filter(estado=False).count()
    total_juridicas  = proveedores.filter(tipo_persona='juridica').count()

    context = {
        'proveedores':      proveedores,
        'fecha_generacion': timezone.now(),
        'total_proveedores': total_proveedores,
        'total_activos':    total_activos,
        'total_inactivos':  total_inactivos,
        'total_juridicas':  total_juridicas,
    }

    template = get_template('reporte_proveedores_pdf.html')
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Proveedores_Huina.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF del reporte de proveedores.', status=500)
    return response


# 7. EXPORTAR PROVEEDORES A EXCEL
@login_required
def export_proveedores_excel(request):
    proveedores = Proveedor.objects.all().order_by('nombre_contacto')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Proveedores_Huina.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proveedores'

    # --- Estilos ---
    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    al_centro  = Alignment(horizontal='center', vertical='center')
    al_derecha = Alignment(horizontal='right',  vertical='center')
    al_izq     = Alignment(horizontal='left',   vertical='center')

    # -- Fila 1: Título --
    ws.append(['REPORTE DE PROVEEDORES - PESCADERÍA HUINA'])
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.font  = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    c.fill  = PatternFill(start_color='0F3976', fill_type='solid')
    c.alignment = al_centro
    ws.row_dimensions[1].height = 26

    # -- Fila 2: Subtítulo --
    ws.append([f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}  |  Total registros: {proveedores.count()}'])
    ws.merge_cells('A2:H2')
    ws['A2'].font      = Font(italic=True, color='6C757D', size=9)
    ws['A2'].alignment = al_centro
    ws.row_dimensions[2].height = 16

    ws.append([])  # Fila 3 vacía

    # -- Fila 4: Cabeceras de tabla --
    encabezados = ['#', 'Tipo Persona', 'NIT / Cédula', 'Nombre / Contacto',
                   'Teléfono', 'Ciudad', 'Estado', 'Fecha Registro']
    ws.append(encabezados)
    for celda in ws[4]:
        celda.font      = Font(bold=True, color='FFFFFF', size=10)
        celda.fill      = PatternFill(start_color='343A40', fill_type='solid')
        celda.alignment = al_centro
        celda.border    = borde
    ws.row_dimensions[4].height = 20

    # -- Filas de datos --
    for idx, p in enumerate(proveedores, start=1):
        fila = idx + 4  # offset por encabezados
        tipo_texto   = 'Natural' if p.tipo_persona == 'natural' else 'Jurídica'
        estado_texto = 'Activo'  if p.estado else 'Inactivo'
        fecha_str    = p.fecha_registro.strftime('%d/%m/%Y') if p.fecha_registro else ''

        ws.append([idx, tipo_texto, p.nit, p.nombre_contacto,
                   p.telefono, p.ciudad, estado_texto, fecha_str])

        for col_idx, celda in enumerate(ws[fila], 1):
            celda.border    = borde
            celda.alignment = al_centro if col_idx in [1, 2, 5, 7, 8] else al_izq

        # Color condicional estado
        celda_estado = ws.cell(row=fila, column=7)
        if estado_texto == 'Activo':
            celda_estado.font = Font(color='065F46', bold=True)
        else:
            celda_estado.font = Font(color='991B1B', bold=True)

    # -- Fila totales --
    fila_total = proveedores.count() + 5
    ws.cell(row=fila_total, column=1, value='TOTAL')
    ws.cell(row=fila_total, column=2, value=proveedores.count())
    for col in range(1, 9):
        c = ws.cell(row=fila_total, column=col)
        c.font   = Font(bold=True, color='0F3976')
        c.border = borde
        c.fill   = PatternFill(start_color='E7F1FF', fill_type='solid')
    ws.cell(row=fila_total, column=1).alignment = al_centro
    ws.cell(row=fila_total, column=2).alignment = al_centro

    # -- Ancho de columnas --
    anchos = {'A': 6, 'B': 14, 'C': 18, 'D': 35, 'E': 16, 'F': 18, 'G': 12, 'H': 15}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    wb.save(response)
    return response


# 5. ELIMINAR PROVEEDOR (SOFT DELETE)
@login_required
def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    
    if request.method == 'POST':
        # Soft delete: Solo cambiamos el estado
        proveedor.estado = False 
        proveedor.save()
        return redirect('proveedores:lista_proveedores')
        
    # CORRECCIÓN AQUÍ
    return render(request, 'eliminar_proveedores.html', {'proveedor': proveedor})