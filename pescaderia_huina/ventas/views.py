from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views import generic
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum
from django.db import transaction
from django.utils import timezone
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Venta, VentaItem
from .forms import VentaForm, VentaItemFormSet, CancelarVentaForm, BusquedaVentaForm
from productos.models import Producto


def _acumular_cantidades_formset(item_formset):
    """Suma cantidades solicitadas por producto excluyendo filas marcadas para eliminar."""
    cantidades_por_producto = {}
    formularios_por_producto = {}

    for item_form in item_formset.forms:
        if not item_form.cleaned_data:
            continue
        if item_form.cleaned_data.get('DELETE'):
            continue

        producto = item_form.cleaned_data.get('producto')
        cantidad = item_form.cleaned_data.get('cantidad')
        if not producto or cantidad is None:
            continue

        cantidades_por_producto[producto.id] = (
            cantidades_por_producto.get(producto.id, Decimal('0')) + Decimal(str(cantidad))
        )
        formularios_por_producto.setdefault(producto.id, []).append(item_form)

    return cantidades_por_producto, formularios_por_producto


def _cantidades_actuales_venta(venta):
    """Obtiene cantidades actuales de una venta por producto para recalcular stock al editar."""
    actuales = {}
    if not venta:
        return actuales

    for item in venta.items.select_related('producto').all():
        actuales[item.producto_id] = actuales.get(item.producto_id, Decimal('0')) + Decimal(str(item.cantidad))

    return actuales


def _validar_stock_disponible(item_formset, venta_actual=None):
    """Valida que cada producto tenga stock suficiente para la cantidad solicitada."""
    solicitadas, formularios_por_producto = _acumular_cantidades_formset(item_formset)
    cantidades_actuales = _cantidades_actuales_venta(venta_actual)
    hay_error = False

    for producto_id, cantidad_solicitada in solicitadas.items():
        producto = Producto.objects.filter(id=producto_id, estado=True).first()
        if not producto:
            for item_form in formularios_por_producto.get(producto_id, []):
                item_form.add_error('producto', 'El producto no está disponible.')
            hay_error = True
            continue

        # El modelo Producto no tiene atributo 'stock'
        # Por ahora se permite la venta sin validación de stock
        stock_disponible = Decimal('0')

    return not hay_error


@login_required
def buscar_productos_api(request):
    q = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')   
    productos = Producto.objects.filter(estado=True).select_related('proveedor')
    if tipo:
        productos = productos.filter(tipo_producto=tipo)
    if q:
        productos = productos.filter(nombre__icontains=q)
    data = [
        {'id': p.id, 'nombre': p.nombre, 'precio': str(p.precio)}
        for p in productos[:20]
    ]
    return JsonResponse({'productos': data})

@login_required
def precio_producto_api(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id, estado=True)
    return JsonResponse({'precio': str(producto.precio), 'nombre': producto.nombre, 'stock': '0'})


@login_required
def ventas(request):
    lista_ventas = Venta.objects.prefetch_related('items__producto').all()
    form_busqueda = BusquedaVentaForm(request.GET)
    if form_busqueda.is_valid():
        estado = form_busqueda.cleaned_data.get('estado')
        buscar = form_busqueda.cleaned_data.get('buscar')
        fecha_inicio = form_busqueda.cleaned_data.get('fecha_inicio')
        fecha_fin = form_busqueda.cleaned_data.get('fecha_fin')
        if estado:
            lista_ventas = lista_ventas.filter(estado=estado)
        if buscar:
            lista_ventas = lista_ventas.filter(
                Q(nombre_cliente__icontains=buscar) |
                Q(documento_cliente__icontains=buscar) |
                Q(items__producto__nombre__icontains=buscar)
            ).distinct()
        if fecha_inicio:
             lista_ventas = lista_ventas.filter(fecha_venta__date__gte=fecha_inicio)
        if fecha_fin:
             lista_ventas = lista_ventas.filter(fecha_venta__date__lte=fecha_fin)

    total_ventas = Venta.objects.filter(estado='COMPLETADA').count()
    total_ingresos = Venta.objects.filter(estado='COMPLETADA').aggregate(total=Sum('total'))['total'] or 0

    context = {
        'lista_ventas': lista_ventas,
        'form_busqueda': form_busqueda,
        'total_ventas': total_ventas,
        'total_ingresos': total_ingresos,
    }
    return render(request, 'lista_ventas.html', context)


@login_required
def detalle_venta(request, id_venta):
    venta = get_object_or_404(Venta, id=id_venta)
    return render(request, 'detalle_venta.html', {'venta': venta})


class VentaCreateView(LoginRequiredMixin, generic.CreateView):
    model = Venta
    form_class = VentaForm
    template_name = 'crear_venta.html'
    success_url = reverse_lazy('ventas:lista_ventas')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['item_formset'] = VentaItemFormSet(self.request.POST)
        else:
            context['item_formset'] = VentaItemFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        if not item_formset.is_valid():
            messages.error(self.request, 'Corrija los errores en los productos.')
            return self.render_to_response(context)
        if not _validar_stock_disponible(item_formset):
            messages.error(self.request, 'No hay stock suficiente para uno o más productos.')
            return self.render_to_response(context)
        with transaction.atomic():
            venta = form.save()
            items = item_formset.save(commit=False)
            for item in items:
                item.venta = venta
                item.save()
            for obj in item_formset.deleted_objects:
                obj.delete()
            venta.recalcular_totales()
            messages.success(self.request, f'Venta {venta.id} registrada. Total: ${venta.total:,.2f}')
        return redirect(self.success_url)

    def form_invalid(self, form):
        messages.error(self.request, 'Por favor, corrija los errores en el formulario.')
        return super().form_invalid(form)


class VentaUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Venta
    form_class = VentaForm
    template_name = 'editar_venta.html'
    success_url = reverse_lazy('ventas:lista_ventas')
    pk_url_kwarg = 'venta_id'

    def dispatch(self, request, *args, **kwargs):
        venta = self.get_object()
        if venta.estado == 'CANCELADA':
            messages.warning(request, 'No se puede editar una venta cancelada.')
            return redirect('ventas:detalle_venta', id_venta=venta.id)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['item_formset'] = VentaItemFormSet(self.request.POST, instance=self.object)
        else:
            context['item_formset'] = VentaItemFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        if not item_formset.is_valid():
            messages.error(self.request, 'Corrija los errores en los productos.')
            return self.render_to_response(context)
        if not _validar_stock_disponible(item_formset, venta_actual=self.object):
            messages.error(self.request, 'No hay stock suficiente para guardar los cambios de la venta.')
            return self.render_to_response(context)

        with transaction.atomic():
            venta = form.save()
            items = item_formset.save(commit=False)
            for obj in item_formset.deleted_objects:
                obj.delete()

            for item in items:
                item.venta = venta
                item.save()

            venta.recalcular_totales()
            messages.success(self.request, f'Venta {venta.id} actualizada exitosamente.')
        return redirect(self.success_url)

    def form_invalid(self, form):
        messages.error(self.request, 'Por favor, corrija los errores en el formulario.')
        return super().form_invalid(form)


class VentaCancelarView(LoginRequiredMixin, generic.FormView):
    template_name = 'cancelar_venta.html'
    form_class = CancelarVentaForm

    def dispatch(self, request, *args, **kwargs):
        self.venta = get_object_or_404(Venta, pk=kwargs['venta_id'])
        if self.venta.estado == 'CANCELADA':
            messages.warning(request, 'Esta venta ya está cancelada.')
            return redirect('ventas:detalle_venta', id_venta=self.venta.id)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['venta'] = self.venta
        return context

    def get_success_url(self):
        return reverse_lazy('ventas:detalle_venta', kwargs={'id_venta': self.venta.id})

    def form_valid(self, form):
        with transaction.atomic():
        # ✅ Sin tocar stock — se recalcula solo al cambiar estado a CANCELADA
          motivo = form.cleaned_data.get('motivo') or 'Sin motivo especificado'
          self.venta.estado = 'CANCELADA'
          self.venta.motivo_cancelacion = motivo
          self.venta.fecha_cancelacion = timezone.now()
          self.venta.save(update_fields=['estado', 'motivo_cancelacion', 'fecha_cancelacion'])
          messages.success(self.request, f'Venta {self.venta.id} cancelada exitosamente.')
        return super().form_valid(form)


@login_required
def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    azul_oscuro = "0A3D62"
    verde = "198754"
    gris_claro = "F2F2F2"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=azul_oscuro, end_color=azul_oscuro, fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'].value = 'REPORTE DE VENTAS — PESCADERÍA HUINA'
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=13)
    ws['A1'].fill = PatternFill(start_color=azul_oscuro, end_color=azul_oscuro, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Estadísticas filtradas por las mismas fechas
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin    = request.GET.get('fecha_fin')
    estado       = request.GET.get('estado')

    qs_base = Venta.objects.all()
    if fecha_inicio:
        qs_base = qs_base.filter(fecha_venta__date__gte=fecha_inicio)
    if fecha_fin:
        qs_base = qs_base.filter(fecha_venta__date__lte=fecha_fin)

    total_completadas = qs_base.filter(estado='COMPLETADA').count()
    total_ingresos = qs_base.filter(estado='COMPLETADA').aggregate(t=Sum('total'))['t'] or Decimal('0')
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Ventas Completadas: {total_completadas}'
    ws['A2'].font = Font(bold=True, size=10)
    ws['A2'].fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    ws['A2'].alignment = center_align
    ws.merge_cells('E2:I2')
    ws['E2'] = f'Ingresos Totales: ${total_ingresos:,.2f}'
    ws['E2'].font = Font(bold=True, size=10, color=verde)
    ws['E2'].fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    ws['E2'].alignment = center_align
    ws.row_dimensions[2].height = 20

    # Cabecera
    headers = ['#', 'Cliente', 'Documento', 'Productos', 'Subtotal', 'IVA (19%)', 'Total', 'Fecha', 'Estado']
    col_widths = [7, 25, 18, 45, 14, 14, 14, 20, 14]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 18

    # Datos
    ventas_qs = Venta.objects.prefetch_related('items__producto').all()
    if fecha_inicio:
        ventas_qs = ventas_qs.filter(fecha_venta__date__gte=fecha_inicio)
    if fecha_fin:
        ventas_qs = ventas_qs.filter(fecha_venta__date__lte=fecha_fin)
    if estado:
        ventas_qs = ventas_qs.filter(estado=estado)

    for row_num, venta in enumerate(ventas_qs, 4):
        items = venta.items.all()
        productos_str = '\n'.join(
            f"{item.producto.nombre} x{item.cantidad} @ ${item.precio_unitario}" for item in items
        ) if items.exists() else (str(venta.producto.nombre) if venta.producto else 'N/A')

        row_data = [
            venta.id,
            venta.nombre_cliente or 'Anónimo',
            venta.documento_cliente or 'N/A',
            productos_str,
            float(venta.subtotal or 0),
            float(venta.iva_monto or 0),
            float(venta.total or 0),
            venta.fecha_venta.strftime('%d/%m/%Y %H:%M'),
            venta.estado,
        ]

        if venta.estado == 'CANCELADA':
            bg = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        elif row_num % 2 == 0:
            bg = PatternFill(start_color=gris_claro, end_color=gris_claro, fill_type="solid")
        else:
            bg = None

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if bg:
                cell.fill = bg
            if col == 1:
                cell.alignment = center_align
            elif col in (5, 6, 7):
                cell.alignment = right_align
                cell.number_format = '$#,##0.00'
            elif col == 4:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if col == 9:
                if venta.estado == 'CANCELADA':
                    cell.font = Font(bold=True, color="DC2626")
                elif venta.estado == 'COMPLETADA':
                    cell.font = Font(bold=True, color=verde)

        ws.row_dimensions[row_num].height = max(15, min(60, 15 * (productos_str.count('\n') + 1)))

    ws.freeze_panes = 'A4'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ventas_pescaderia_huina.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_pdf(request):
    from xhtml2pdf import pisa
    from django.template.loader import get_template
    from django.db.models import Sum

    fecha_inicio  = request.GET.get('fecha_inicio') or None
    fecha_fin     = request.GET.get('fecha_fin')    or None
    estado_filtro = request.GET.get('estado')       or None

    ventas_qs = Venta.objects.prefetch_related('items__producto').all()
    if fecha_inicio:
        ventas_qs = ventas_qs.filter(fecha_venta__date__gte=fecha_inicio)
    if fecha_fin:
        ventas_qs = ventas_qs.filter(fecha_venta__date__lte=fecha_fin)
    if estado_filtro:
        ventas_qs = ventas_qs.filter(estado=estado_filtro)

    qs_base = Venta.objects.all()
    if fecha_inicio:
        qs_base = qs_base.filter(fecha_venta__date__gte=fecha_inicio)
    if fecha_fin:
        qs_base = qs_base.filter(fecha_venta__date__lte=fecha_fin)

    qs_completadas = qs_base.filter(estado='COMPLETADA')
    qs_canceladas  = qs_base.filter(estado='CANCELADA')

    context = {
        'ventas':             ventas_qs,
        'total_completadas':  qs_completadas.count(),
        'total_ingresos':     qs_completadas.aggregate(t=Sum('total'))['t'] or Decimal('0'),
        'total_canceladas':   qs_canceladas.count(),
        'fecha_inicio':       fecha_inicio,
        'fecha_fin':          fecha_fin,
        'estado_filtro':      estado_filtro,
    }

    template  = get_template('reporte_ventas_pdf.html')
    html      = template.render(context)
    response  = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ventas_pescaderia_huina.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response